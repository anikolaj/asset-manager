import math
from openpyxl import styles, Workbook, utils, chart
from typing import Optional

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fills import PatternFill

import asset_manager.math_functions as mf
from asset_manager.objects import Interval
from asset_manager.portfolio_analyzer import PortfolioAnalyzer

YELLOW_BACKGROUND = PatternFill(fgColor=styles.colors.COLOR_INDEX[5], fill_type="solid")


class ExcelUtility:

    def __init__(self, portfolio_analyzer: PortfolioAnalyzer) -> None:
        self.portfolio_analyzer = portfolio_analyzer

        self.workbook = Workbook()
        self.file_name = self.portfolio_analyzer.portfolio.name + "_analysis.xlsx"

        self.monthly_expected_return_row = None
        self.monthly_expected_return_column = None

        self.monthly_standard_deviation_row = None
        self.monthly_standard_deviation_column = None

    def generate_portfolio_workbook(self) -> None:
        self.write_summary()
        self.write_statistics()
        self.write_mvp()
        self.write_mvl()

        self.workbook.save(filename=self.file_name)

    def write_summary(self) -> None:
        summary_sheet = self.workbook.active
        summary_sheet.title = "Summary"

        sheet_title = f"Portfolio Name - {self.portfolio_analyzer.portfolio.name}"
        self.set_cell(summary_sheet, summary_sheet["A1"], sheet_title, True, YELLOW_BACKGROUND)

        # write asset details
        current_row = 3
        self.write_asset_summary(current_row, summary_sheet)

        # write portfolio summary
        current_row += len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_portfolio_summary(current_row, summary_sheet)

        # create border layout for summary sheet
        self.create_summary_border(summary_sheet)

    def write_asset_summary(self, current_row: int, summary_sheet: Worksheet) -> None:
        # set the asset summary header
        self.set_cell(summary_sheet, summary_sheet["A2"], "Asset Summary", True, None)
        self.set_cell(summary_sheet, summary_sheet["B2"], "Current", True, None)
        self.set_cell(summary_sheet, summary_sheet["C2"], "Ticker", True, None)
        self.set_cell(summary_sheet, summary_sheet["D2"], "Price", True, None)
        self.set_cell(summary_sheet, summary_sheet["E2"], "Shares", True, None)
        self.set_cell(summary_sheet, summary_sheet["F2"], "Total Value", True, None)
        self.set_cell(summary_sheet, summary_sheet["G2"], "Weight", True, None)

        current_row = 3

        for equity in self.portfolio_analyzer.portfolio.equities:
            # ticker
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=3), equity.ticker, False, None)

            # price
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=4), equity.price, False, None)

            # shares
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=5), equity.shares, False, None)

            # total value
            total_value = equity.price * equity.shares
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=6), total_value, False, None)

            # weight
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=7), equity.weight, False, None)

            current_row += 1

    def write_portfolio_summary(self, current_row: int, summary_sheet: Worksheet) -> None:
        # section summary
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=1), "Portfolio Summary", True, None)

        # column headers
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=3), "Monthly (1M)", True, None)
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=4), "Quarterly (1Q)", True, None)
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=5), "Biannually (2Q)", True, None)
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=6), "Yearly (1Y)", True, None)
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=7), "5-Year (5Y)", True, None)

        current_row += 1

        # expected return
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=2), "Expected Return", False, None)

        start_column = 3
        for time_interval in [Interval.MONTH, Interval.THREE_MONTH, Interval.SIX_MONTH, Interval.YEAR, Interval.FIVE_YEAR]:
            expected_return = self.portfolio_analyzer.expected_return[time_interval]
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=start_column), expected_return, False, None)
            if time_interval == Interval.YEAR:
                self.monthly_expected_return_row = current_row
                self.monthly_expected_return_column = start_column

            start_column += 1

        current_row += 1

        # standard deviation
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=2), "Standard Deviation", False, None)

        start_column = 3
        for time_interval in [Interval.MONTH, Interval.THREE_MONTH, Interval.SIX_MONTH, Interval.YEAR, Interval.FIVE_YEAR]:
            standard_deviation = self.portfolio_analyzer.standard_deviation[time_interval]
            self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=start_column), standard_deviation, False, None)
            if time_interval == Interval.YEAR:
                self.monthly_standard_deviation_row = current_row
                self.monthly_standard_deviation_column = start_column

            start_column += 1

        current_row += 2

        # portfolio value
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=1), "Total Value", True, YELLOW_BACKGROUND)

        portfolio_value = self.portfolio_analyzer.portfolio.value
        self.set_cell(summary_sheet, summary_sheet.cell(row=current_row, column=2), portfolio_value, True, YELLOW_BACKGROUND)

    def create_summary_border(self, summary_sheet: Worksheet) -> None:
        num_columns = 8
        cell_length = len(self.portfolio_analyzer.portfolio.equities) + 8

        # column A line for asset summary
        for i in range(1, cell_length):
            summary_sheet.cell(row=i, column=1).border = styles.borders.Border(right=styles.borders.Side(style="thin"))

        # row 2 line for asset summary
        for i in range(1, num_columns):
            cell = summary_sheet.cell(row=2, column=i)

            if cell.border.right.style == "thin":
                cell.border = styles.borders.Border(right=styles.borders.Side(style="thin"), bottom=styles.borders.Side(style="thin"))
            else:
                cell.border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

        # column G line for both summaries
        for i in range(3, cell_length):
            summary_sheet.cell(row=i, column=7).border = styles.borders.Border(right=styles.borders.Side(style="thin"))

        # row 7 and row 10 line for portfolio summary
        portfolio_summary_start_row = len(self.portfolio_analyzer.portfolio.equities) + 5
        for i in range(1, num_columns):
            self.set_cell_border(summary_sheet.cell(row=portfolio_summary_start_row - 1, column=i))
            self.set_cell_border(summary_sheet.cell(row=cell_length - 1, column=i))

    def set_cell_border(self, cell: Cell) -> None:
        if cell.border.right.style == "thin":
            cell.border = styles.borders.Border(right=styles.borders.Side(style="thin"), bottom=styles.borders.Side(style="thin"))
        else:
            cell.border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

    def write_statistics(self) -> None:
        stats_sheet = self.workbook.create_sheet("Statistics")

        self.set_cell(stats_sheet, stats_sheet["A1"], "Portfolio Statistics", True, YELLOW_BACKGROUND)

        # monthly statistics
        start_row = 2
        self.write_time_statistics(stats_sheet, start_row, Interval.MONTH)

        # quarterly statistics
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_statistics(stats_sheet, start_row, Interval.THREE_MONTH)

        # biannual statistics
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_statistics(stats_sheet, start_row, Interval.SIX_MONTH)

        # yearly statistics
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_statistics(stats_sheet, start_row, Interval.YEAR)

        # 5 year statistics
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_statistics(stats_sheet, start_row, Interval.FIVE_YEAR)

    def write_mvp(self) -> None:
        mvp_sheet = self.workbook.create_sheet("MVP")

        self.set_cell(mvp_sheet, mvp_sheet["A1"], "Minimum Variance Portfolio", True, YELLOW_BACKGROUND)

        # monthly minimum variance portfolio
        start_row = 2
        self.write_time_mvp(mvp_sheet, start_row, Interval.MONTH)

        # quarterly minimum variance portfolio
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_mvp(mvp_sheet, start_row, Interval.THREE_MONTH)

        # biannual minimum variance portfolio
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_mvp(mvp_sheet, start_row, Interval.SIX_MONTH)

        # yearly minimum variance portfolio
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_mvp(mvp_sheet, start_row, Interval.YEAR)

        # 5 year minimum variance portfolio
        start_row = start_row + len(self.portfolio_analyzer.portfolio.equities) + 2
        self.write_time_mvp(mvp_sheet, start_row, Interval.FIVE_YEAR)

    def write_mvl(self) -> None:
        mvl_sheet = self.workbook.create_sheet("MVL")

        self.set_cell(mvl_sheet, mvl_sheet["A1"], "Minimum Variance Line - Monthly", True, YELLOW_BACKGROUND)

        self.set_cell(mvl_sheet, mvl_sheet["A2"], "Standard Deviation", False, None)
        mvl_sheet["A2"].border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

        self.set_cell(mvl_sheet, mvl_sheet["B2"], "Efficient Expected Return", False, None)
        mvl_sheet["B2"].border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

        self.set_cell(mvl_sheet, mvl_sheet["C2"], "Inefficient Expected Return", False, None)
        mvl_sheet["C2"].border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

        self.write_time_mvl(mvl_sheet, 3, Interval.YEAR)

    def resize_column(self, worksheet: Worksheet, row: int, column: int) -> None:
        column_str = utils.get_column_letter(column)
        # this line is set because formating with a specified number format takes the original value not the formatted value
        if worksheet.cell(row, column).number_format == "General":
            if worksheet.column_dimensions[column_str].width < len(str(worksheet.cell(row, column).value)):
                worksheet.column_dimensions[column_str].width = len(str(worksheet.cell(row, column).value))

    def write_time_statistics(self, worksheet: Worksheet, current_row: int, time_interval: Interval) -> None:
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=2), time_interval.value, True, None)

        current_row += 1

        self.write_ticker_statistics(worksheet, current_row, time_interval)
        self.write_correlation_matrix(worksheet, current_row, time_interval)

    def write_ticker_statistics(self, worksheet: Worksheet, current_row: int, time_interval: Interval) -> None:
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=3), "Ticker", True, None)
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=4), "Expected Return", True, None)
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=5), "Standard Deviation", True, None)

        ticker_row = current_row + 1

        for eq in self.portfolio_analyzer.portfolio.equities:
            # ticker
            self.set_cell(worksheet, worksheet.cell(row=ticker_row, column=3), eq.ticker, False, None)

            # expected return
            expected_return = self.portfolio_analyzer.ticker_to_timeseries[eq.ticker][time_interval].avg_return
            self.set_cell(worksheet, worksheet.cell(row=ticker_row, column=4), expected_return, False, None, "0.0000000000")

            # standard deviation
            std_dev = self.portfolio_analyzer.ticker_to_timeseries[eq.ticker][time_interval].std_dev
            self.set_cell(worksheet, worksheet.cell(row=ticker_row, column=5), std_dev, False, None, "0.0000000000")

            ticker_row += 1

    def write_correlation_matrix(self, worksheet: Worksheet, current_row: int, time_interval: Interval) -> None:
        if self.portfolio_analyzer.C[time_interval].size == 0:
            current_row += len(self.portfolio_analyzer.portfolio.equities)
            return

        start_column = 7
        current_column = start_column

        self.set_cell(worksheet, worksheet.cell(row=current_row, column=current_column), "Correlation Matrix", True, None)

        # write top row of correlation matrix
        current_column += 2

        for eq in self.portfolio_analyzer.portfolio.equities:
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=current_column), eq.ticker, False, None)
            worksheet.cell(row=current_row, column=current_column).border = styles.borders.Border(bottom=styles.borders.Side(style="thin"))

            current_column += 1

        # write column and values of correlation matrix
        current_column = start_column + 1

        for i in range(0, len(self.portfolio_analyzer.portfolio.equities)):
            current_row += 1
            current_equity = self.portfolio_analyzer.portfolio.equities[i]

            self.set_cell(worksheet, worksheet.cell(row=current_row, column=current_column), current_equity.ticker, False, None)
            worksheet.cell(row=current_row, column=current_column).border = styles.borders.Border(right=styles.borders.Side(style="thin"))

            # write correlation coefficients into matrix
            for j in range(0, len(self.portfolio_analyzer.portfolio.equities)):
                current_column += 1
                correlation_coefficient = self.portfolio_analyzer.C[time_interval][i, j]
                self.set_cell(worksheet, worksheet.cell(row=current_row, column=current_column), correlation_coefficient, False, None, "0.00000")

                if j == (len(self.portfolio_analyzer.portfolio.equities) - 1):
                    worksheet.cell(row=current_row, column=current_column + 1).border = styles.borders.Border(left=styles.borders.Side(style="thin"))

                if i == (len(self.portfolio_analyzer.portfolio.equities) - 1):
                    worksheet.cell(row=current_row + 1, column=current_column).border = styles.borders.Border(top=styles.borders.Side(style="thin"))

            current_column = start_column + 1

    def write_time_mvp(self, worksheet: Worksheet, current_row: int, time_interval: Interval) -> None:
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=2), time_interval.value, True, None)

        current_row += 1

        self.write_ticker_mvp(worksheet, current_row, time_interval)

    def write_ticker_mvp(self, worksheet: Worksheet, current_row: int, time_interval: Interval) -> None:
        if len(self.portfolio_analyzer.mvp[time_interval]) == 0:
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=3), "Not Applied", True, None)
            return

        self.set_cell(worksheet, worksheet.cell(row=current_row, column=3), "Ticker", True, None)
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=4), "Weight", True, None)
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=5), "Shares", True, None)
        self.set_cell(worksheet, worksheet.cell(row=current_row, column=6), "Value Change", True, None)

        for i in range(0, len(self.portfolio_analyzer.portfolio.equities)):
            current_row += 1
            current_equity = self.portfolio_analyzer.portfolio.equities[i]

            # ticker
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=3), current_equity.ticker, False, None)

            # weight
            mvp_weight = self.portfolio_analyzer.mvp[time_interval][i]
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=4), mvp_weight, False, None)

            # shares
            mvp_shares = (mvp_weight * self.portfolio_analyzer.portfolio.value) / current_equity.price
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=5), mvp_shares, False, None)

            # value change
            new_value = mvp_shares * current_equity.price
            old_value = current_equity.shares * current_equity.price
            value_change = new_value - old_value
            self.set_cell(worksheet, worksheet.cell(row=current_row, column=6), value_change, False, None)

    def write_time_mvl(self, worksheet: Worksheet, current_row: int, time_interval: str) -> None:
        mvp_expected_return = mf.calculate_expected_value(self.portfolio_analyzer.M[time_interval], self.portfolio_analyzer.mvp[time_interval])
        mvp_std_dev = round(math.sqrt(mf.calculate_variance(self.portfolio_analyzer.C[time_interval], self.portfolio_analyzer.mvp[time_interval])), 8)

        self.set_mvp_entry(worksheet, current_row, mvp_std_dev, mvp_expected_return, mvp_expected_return)

        row = current_row + 1

        for i in range(1, 50, 2):
            delta_v = i / 100
            sigma_v = mvp_std_dev + delta_v

            m_v1, m_v2 = mf.solve_quadratic_formula(self.portfolio_analyzer.mvl_a[time_interval], self.portfolio_analyzer.mvl_b[time_interval], self.portfolio_analyzer.mvl_c[time_interval] - (sigma_v ** 2))

            self.set_mvp_entry(worksheet, row, sigma_v, m_v1, m_v2)

            row = row + 1

        mvl_chart = chart.ScatterChart()

        std_devs = chart.Reference(worksheet, min_col=1, min_row=3, max_col=1, max_row=28)
        efficient_returns = chart.Reference(worksheet, min_col=2, min_row=3, max_col=2, max_row=28)
        series1 = chart.Series(efficient_returns, std_devs, title="Efficient Frontier")

        series1.marker = chart.marker.Marker('circle')
        # series1.graphicalProperties.line.noFill=True
        mvl_chart.series.append(series1)

        inefficient_returns = chart.Reference(worksheet, min_col=3, min_row=3, max_col=3, max_row=28)
        series2 = chart.Series(inefficient_returns, std_devs, title="Inefficient Frontier")

        series2.marker = chart.marker.Marker('circle')
        # series2.graphicalProperties.line.noFill=True
        mvl_chart.series.append(series2)

        current_std_dev = chart.Reference(self.workbook["Summary"], min_col=self.monthly_standard_deviation_column, min_row=self.monthly_standard_deviation_row, max_col=self.monthly_standard_deviation_column, max_row=self.monthly_standard_deviation_row)
        current_expected_return = chart.Reference(self.workbook["Summary"], min_col=self.monthly_expected_return_column, min_row=self.monthly_expected_return_row, max_col=self.monthly_expected_return_column, max_row=self.monthly_expected_return_row)
        series3 = chart.Series(current_expected_return, current_std_dev, title="Current Allocation")

        series3.marker = chart.marker.Marker('circle')
        # series2.graphicalProperties.line.noFill=True
        mvl_chart.series.append(series3)

        mvl_chart.title = "Minimum Variance Line"
        mvl_chart.x_axis.title = "Standard Deviation"
        mvl_chart.y_axis.title = "Expected Return"
        mvl_chart.height = 10
        mvl_chart.width = 20

        worksheet.add_chart(mvl_chart, "E3")

    def set_mvp_entry(self, worksheet: Worksheet, row: int, sigma_v: float, m_v1: float, m_v2: float) -> None:
        self.set_cell(worksheet, worksheet.cell(row=row, column=1), sigma_v, False, None)
        self.set_cell(worksheet, worksheet.cell(row=row, column=2), m_v1, False, None)
        self.set_cell(worksheet, worksheet.cell(row=row, column=3), m_v2, False, None)
        worksheet.cell(row=row, column=3).border = styles.borders.Border(right=styles.borders.Side(style="thin"))

    def set_cell(
        self,
        sheet: Worksheet,
        cell: Cell,
        value: str,
        bold: bool,
        fill: PatternFill,
        number_format: Optional[str] = None
    ) -> None:
        cell.value = value

        if bold is True:
            cell.font = styles.Font(bold=True)

        if fill is not None:
            cell.fill = fill

        if number_format is not None:
            cell.number_format = number_format

        self.resize_column(sheet, cell.row, cell.column)
